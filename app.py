from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from datetime import datetime
from database import db
from models import *
import openpyxl

app = Flask(__name__)
import os
database_url = os.environ.get('DATABASE_URL', 'sqlite:///financeiro.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'sua-chave-secreta-aqui'

db.init_app(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Faça login para acessar o sistema.'

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

@app.before_request
def criar_tabelas():
    db.create_all()

@app.template_filter('moeda')
def moeda_filter(value):
    if value is None:
        return 'R$ 0,00'
    return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

MESES = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março',
    4: 'Abril', 5: 'Maio', 6: 'Junho',
    7: 'Julho', 8: 'Agosto', 9: 'Setembro',
    10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

@app.context_processor
def inject_now():
    hoje = datetime.now()
    mes_nome = MESES[hoje.month]
    return {
        'now': hoje,
        'mes_atual': f"{mes_nome}/{hoje.year}"
    }

# AUTH
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('painel'))
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        usuario = Usuario.query.filter_by(email=email).first()
        if usuario and bcrypt.check_password_hash(usuario.senha, senha):
            if not usuario.ativo:
                flash('Conta desativada. Entre em contato com o administrador.', 'erro')
                return redirect(url_for('login'))
            login_user(usuario)
            return redirect(url_for('painel'))
        flash('Email ou senha incorretos.', 'erro')
    return render_template('login.html')

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    if current_user.is_authenticated:
        return redirect(url_for('painel'))
    if request.method == 'POST':
        nome = request.form['nome']
        email = request.form['email']
        senha = request.form['senha']
        if Usuario.query.filter_by(email=email).first():
            flash('Este email já está cadastrado.', 'erro')
            return redirect(url_for('cadastro'))
        senha_hash = bcrypt.generate_password_hash(senha).decode('utf-8')
        usuario = Usuario(nome=nome, email=email, senha=senha_hash)
        db.session.add(usuario)
        db.session.commit()
        login_user(usuario)
        return redirect(url_for('painel'))
    return render_template('cadastro.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ADMIN
@app.route('/admin')
@login_required
def admin():
    if not current_user.admin:
        return f"Acesso negado. Seu admin é: {current_user.admin} | Email: {current_user.email}"
    usuarios = Usuario.query.order_by(Usuario.criado_em.desc()).all()
    for u in usuarios:
        u.total_clientes = Cliente.query.filter_by(usuario_id=u.id).count()
        u.total_despesas = DespesaFixa.query.filter_by(usuario_id=u.id).count()
    return render_template('admin.html', usuarios=usuarios)

@app.route('/admin/toggle/<int:id>', methods=['POST'])
@login_required
def admin_toggle_usuario(id):
    if not current_user.admin:
        return redirect(url_for('painel'))
    usuario = Usuario.query.get_or_404(id)
    usuario.ativo = not usuario.ativo
    db.session.commit()
    return redirect(url_for('admin'))

# PAINEL
@app.route('/')
@login_required
def painel():
    hoje = datetime.now()
    mes = hoje.month
    ano = hoje.year
    uid = current_user.id

    fixas = DespesaFixa.query.filter_by(ativa=True, usuario_id=uid).all()
    for despesa in fixas:
        despesa.pagamento_mes = PagamentoMes.query.filter_by(
            tipo='fixa', referencia_id=despesa.id, mes=mes, ano=ano, pago=True
        ).first()
    fixas = sorted(fixas, key=lambda d: (0 if d.pagamento_mes else 1))

    variaveis = DespesaVariavel.query.filter_by(usuario_id=uid).all()
    for despesa in variaveis:
        despesa.lancamento_mes = LancamentoDespesaVariavel.query.filter_by(
            despesa_id=despesa.id, mes=mes, ano=ano
        ).first()
    variaveis = sorted(variaveis, key=lambda d: (0 if d.lancamento_mes and d.lancamento_mes.pago else 1))

    clientes = Cliente.query.filter_by(status='ativo', usuario_id=uid).all()
    for cliente in clientes:
        cliente.pagamento_mes = PagamentoCliente.query.filter_by(
            cliente_id=cliente.id, mes=mes, ano=ano, pago=True
        ).first()
    clientes = sorted(clientes, key=lambda c: (0 if c.pagamento_mes else 1))

    receitas_fixas = ReceitaPessoal.query.filter_by(fixa=True, usuario_id=uid).all()
    receitas_extras = ReceitaPessoal.query.filter_by(fixa=False, usuario_id=uid).all()
    for extra in receitas_extras:
        extra.lancado_mes = LancamentoReceitaExtra.query.filter_by(
            receita_id=extra.id, mes=mes, ano=ano
        ).first()

    total_receita_fixa = sum(r.valor for r in receitas_fixas)
    total_receita_extra = sum(e.lancado_mes.valor for e in receitas_extras if e.lancado_mes)
    total_clientes = sum(c.valor_mensalidade for c in clientes)
    total_receitas = total_receita_fixa + total_receita_extra + total_clientes
    total_despesas_fixas = sum(d.valor for d in fixas)
    total_despesas_variaveis = sum(d.lancamento_mes.valor for d in variaveis if d.lancamento_mes)
    total_despesas = total_despesas_fixas + total_despesas_variaveis
    saldo_previsto = total_receitas - total_despesas

    contas_abertas = []
    for d in fixas:
        if not d.pagamento_mes:
            contas_abertas.append({'nome': d.descricao, 'valor': d.valor, 'tipo': 'Fixa'})
    for d in variaveis:
        if d.lancamento_mes and not d.lancamento_mes.pago:
            contas_abertas.append({'nome': d.descricao, 'valor': d.lancamento_mes.valor, 'tipo': 'Variável'})

    return render_template('painel.html',
        mes=MESES[mes], ano=ano,
        total_receitas=total_receitas,
        total_despesas=total_despesas,
        saldo_previsto=saldo_previsto,
        contas_abertas=contas_abertas,
        fixas=fixas,
        variaveis=variaveis,
        clientes=clientes
    )

# CLIENTES
@app.route('/clientes')
@login_required
def clientes():
    hoje = datetime.now()
    todos = Cliente.query.filter_by(usuario_id=current_user.id).order_by(Cliente.nome).all()
    for cliente in todos:
        cliente.pagamento_mes = PagamentoCliente.query.filter_by(
            cliente_id=cliente.id, mes=hoje.month, ano=hoje.year, pago=True
        ).first()
    return render_template('clientes.html', clientes=todos)

@app.route('/clientes/novo', methods=['POST'])
@login_required
def novo_cliente():
    nome = request.form['nome']
    valor = float(request.form['valor'])
    dia = int(request.form['dia_vencimento'])
    cliente = Cliente(nome=nome, valor_mensalidade=valor, dia_vencimento=dia, usuario_id=current_user.id)
    db.session.add(cliente)
    db.session.commit()
    return redirect(url_for('clientes'))

@app.route('/clientes/status/<int:id>', methods=['POST'])
@login_required
def atualizar_status_cliente(id):
    cliente = Cliente.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    cliente.status = request.form['status']
    db.session.commit()
    return redirect(url_for('clientes'))

@app.route('/clientes/pagar/<int:id>', methods=['POST'])
@login_required
def pagar_cliente(id):
    hoje = datetime.now()
    cliente = Cliente.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    pagamento = PagamentoCliente.query.filter_by(
        cliente_id=id, mes=hoje.month, ano=hoje.year
    ).first()
    if pagamento:
        pagamento.pago = not pagamento.pago
    else:
        pagamento = PagamentoCliente(cliente_id=id, mes=hoje.month, ano=hoje.year, pago=True)
        db.session.add(pagamento)
    db.session.commit()
    return redirect(url_for('clientes'))

# DESPESAS FIXAS
@app.route('/despesas-fixas')
@login_required
def despesas_fixas():
    hoje = datetime.now()
    despesas = DespesaFixa.query.filter_by(ativa=True, usuario_id=current_user.id).order_by(DespesaFixa.descricao).all()
    for despesa in despesas:
        despesa.pagamento_mes = PagamentoMes.query.filter_by(
            tipo='fixa', referencia_id=despesa.id, mes=hoje.month, ano=hoje.year, pago=True
        ).first()
    return render_template('despesas_fixas.html', despesas=despesas)

@app.route('/despesas-fixas/nova', methods=['POST'])
@login_required
def nova_despesa_fixa():
    descricao = request.form['descricao']
    valor = float(request.form['valor'])
    despesa = DespesaFixa(descricao=descricao, valor=valor, usuario_id=current_user.id)
    db.session.add(despesa)
    db.session.commit()
    return redirect(url_for('despesas_fixas'))

@app.route('/despesas-fixas/pagar/<int:id>', methods=['POST'])
@login_required
def pagar_despesa_fixa(id):
    hoje = datetime.now()
    DespesaFixa.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    pagamento = PagamentoMes.query.filter_by(
        tipo='fixa', referencia_id=id, mes=hoje.month, ano=hoje.year
    ).first()
    if pagamento:
        pagamento.pago = not pagamento.pago
    else:
        pagamento = PagamentoMes(tipo='fixa', referencia_id=id, mes=hoje.month, ano=hoje.year, pago=True)
        db.session.add(pagamento)
    db.session.commit()
    return redirect(url_for('despesas_fixas'))

@app.route('/despesas-fixas/editar/<int:id>', methods=['POST'])
@login_required
def editar_despesa_fixa(id):
    despesa = DespesaFixa.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    novo_valor = float(request.form['valor'])
    if novo_valor != despesa.valor:
        historico = HistoricoReajuste(despesa_id=despesa.id, valor_antigo=despesa.valor, valor_novo=novo_valor)
        db.session.add(historico)
        despesa.valor = novo_valor
    despesa.descricao = request.form['descricao']
    db.session.commit()
    return redirect(url_for('despesas_fixas'))

@app.route('/despesas-fixas/desativar/<int:id>', methods=['POST'])
@login_required
def desativar_despesa_fixa(id):
    despesa = DespesaFixa.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    despesa.ativa = False
    db.session.commit()
    return redirect(url_for('despesas_fixas'))

# DESPESAS VARIÁVEIS
@app.route('/despesas-variaveis')
@login_required
def despesas_variaveis():
    hoje = datetime.now()
    despesas = DespesaVariavel.query.filter_by(usuario_id=current_user.id).order_by(DespesaVariavel.descricao).all()
    for despesa in despesas:
        despesa.lancamento_mes = LancamentoDespesaVariavel.query.filter_by(
            despesa_id=despesa.id, mes=hoje.month, ano=hoje.year
        ).first()
    return render_template('despesas_variaveis.html', despesas=despesas, mes=MESES[hoje.month], ano=hoje.year)

@app.route('/despesas-variaveis/nova', methods=['POST'])
@login_required
def nova_despesa_variavel():
    descricao = request.form['descricao']
    cor = request.form['cor']
    despesa = DespesaVariavel(descricao=descricao, cor=cor, usuario_id=current_user.id)
    db.session.add(despesa)
    db.session.commit()
    return redirect(url_for('despesas_variaveis'))

@app.route('/despesas-variaveis/lancar/<int:id>', methods=['POST'])
@login_required
def lancar_despesa_variavel(id):
    hoje = datetime.now()
    DespesaVariavel.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    valor = float(request.form['valor'])
    lancamento = LancamentoDespesaVariavel.query.filter_by(
        despesa_id=id, mes=hoje.month, ano=hoje.year
    ).first()
    if lancamento:
        lancamento.valor = valor
    else:
        lancamento = LancamentoDespesaVariavel(despesa_id=id, mes=hoje.month, ano=hoje.year, valor=valor, pago=False)
        db.session.add(lancamento)
    db.session.commit()
    return redirect(url_for('despesas_variaveis'))

@app.route('/despesas-variaveis/pagar/<int:id>', methods=['POST'])
@login_required
def pagar_despesa_variavel(id):
    hoje = datetime.now()
    DespesaVariavel.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    lancamento = LancamentoDespesaVariavel.query.filter_by(
        despesa_id=id, mes=hoje.month, ano=hoje.year
    ).first()
    if lancamento:
        lancamento.pago = not lancamento.pago
        db.session.commit()
        if lancamento.pago:
            parcelas = CompraParcelada.query.filter_by(despesa_variavel_id=id, usuario_id=current_user.id).all()
            for compra in parcelas:
                parcela_mes = ParcelaMes.query.filter_by(
                    compra_id=compra.id, mes=hoje.month, ano=hoje.year
                ).first()
                if not parcela_mes:
                    parcela_mes = ParcelaMes(compra_id=compra.id, mes=hoje.month, ano=hoje.year, pago=True)
                    db.session.add(parcela_mes)
                else:
                    parcela_mes.pago = True
            db.session.commit()
    return redirect(url_for('despesas_variaveis'))

@app.route('/despesas-variaveis/editar/<int:id>', methods=['POST'])
@login_required
def editar_despesa_variavel(id):
    despesa = DespesaVariavel.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    despesa.descricao = request.form['descricao']
    db.session.commit()
    return redirect(url_for('despesas_variaveis'))

@app.route('/despesas-variaveis/cor/<int:id>', methods=['POST'])
@login_required
def editar_cor_despesa_variavel(id):
    despesa = DespesaVariavel.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    despesa.cor = request.form['cor']
    db.session.commit()
    return redirect(url_for('despesas_variaveis'))

@app.route('/despesas-variaveis/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_despesa_variavel(id):
    despesa = DespesaVariavel.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    LancamentoDespesaVariavel.query.filter_by(despesa_id=id).delete()
    CompraParcelada.query.filter_by(despesa_variavel_id=id).delete()
    db.session.delete(despesa)
    db.session.commit()
    return redirect(url_for('despesas_variaveis'))

# PARCELAS
@app.route('/parcelas')
@login_required
def parcelas():
    hoje = datetime.now()
    todas = CompraParcelada.query.filter_by(usuario_id=current_user.id).order_by(CompraParcelada.descricao).all()
    despesas_variaveis = DespesaVariavel.query.filter_by(usuario_id=current_user.id).order_by(DespesaVariavel.descricao).all()
    total_restante = 0
    for compra in todas:
        pagas = ParcelaMes.query.filter_by(compra_id=compra.id, pago=True).count()
        compra.parcelas_pagas = pagas
        compra.parcelas_restantes = compra.num_parcelas - pagas
        compra.valor_restante = compra.parcelas_restantes * compra.valor_parcela
        total_restante += compra.valor_restante
        compra.cartao = DespesaVariavel.query.get(compra.despesa_variavel_id) if compra.despesa_variavel_id else None
    return render_template('parcelas.html',
        parcelas=todas,
        despesas_variaveis=despesas_variaveis,
        total_restante=total_restante,
        mes=MESES[hoje.month],
        ano=hoje.year
    )

@app.route('/parcelas/nova', methods=['POST'])
@login_required
def nova_parcela():
    compra = CompraParcelada(
        descricao=request.form['descricao'],
        valor_parcela=float(request.form['valor_parcela']),
        num_parcelas=int(request.form['num_parcelas']),
        mes_inicio=int(request.form['mes_inicio']),
        ano_inicio=int(request.form['ano_inicio']),
        despesa_variavel_id=int(request.form['despesa_variavel_id']),
        usuario_id=current_user.id
    )
    db.session.add(compra)
    db.session.commit()
    return redirect(url_for('parcelas'))

@app.route('/parcelas/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_parcela(id):
    compra = CompraParcelada.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    ParcelaMes.query.filter_by(compra_id=id).delete()
    db.session.delete(compra)
    db.session.commit()
    return redirect(url_for('parcelas'))

# RECEITAS
@app.route('/receitas')
@login_required
def receitas():
    hoje = datetime.now()
    uid = current_user.id
    fixas = ReceitaPessoal.query.filter_by(fixa=True, usuario_id=uid).order_by(ReceitaPessoal.descricao).all()
    extras = ReceitaPessoal.query.filter_by(fixa=False, usuario_id=uid).all()
    for extra in extras:
        extra.lancado_mes = LancamentoReceitaExtra.query.filter_by(
            receita_id=extra.id, mes=hoje.month, ano=hoje.year
        ).first()
    clientes_ativos = Cliente.query.filter_by(status='ativo', usuario_id=uid).all()
    for cliente in clientes_ativos:
        cliente.pagamento_mes = PagamentoCliente.query.filter_by(
            cliente_id=cliente.id, mes=hoje.month, ano=hoje.year, pago=True
        ).first()
    total_fixo = sum(r.valor for r in fixas)
    total_extra = sum(e.lancado_mes.valor for e in extras if e.lancado_mes)
    total_clientes = sum(c.valor_mensalidade for c in clientes_ativos)
    return render_template('receitas.html',
        fixas=fixas, extras=extras, clientes=clientes_ativos,
        total_fixo=total_fixo, total_extra=total_extra,
        total_clientes=total_clientes,
        total_geral=total_fixo + total_extra + total_clientes,
        mes=MESES[hoje.month], ano=hoje.year
    )

@app.route('/receitas/nova-fixa', methods=['POST'])
@login_required
def nova_receita_fixa():
    receita = ReceitaPessoal(
        descricao=request.form['descricao'],
        valor=float(request.form['valor']),
        fixa=True, usuario_id=current_user.id
    )
    db.session.add(receita)
    db.session.commit()
    return redirect(url_for('receitas'))

@app.route('/receitas/nova-extra', methods=['POST'])
@login_required
def nova_receita_extra():
    receita = ReceitaPessoal(
        descricao=request.form['descricao'],
        valor=0, fixa=False, usuario_id=current_user.id
    )
    db.session.add(receita)
    db.session.commit()
    return redirect(url_for('receitas'))

@app.route('/receitas/lancar-extra/<int:id>', methods=['POST'])
@login_required
def lancar_receita_extra(id):
    hoje = datetime.now()
    ReceitaPessoal.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    valor = float(request.form['valor'])
    lancamento = LancamentoReceitaExtra.query.filter_by(
        receita_id=id, mes=hoje.month, ano=hoje.year
    ).first()
    if lancamento:
        lancamento.valor = valor
    else:
        lancamento = LancamentoReceitaExtra(receita_id=id, mes=hoje.month, ano=hoje.year, valor=valor)
        db.session.add(lancamento)
    db.session.commit()
    return redirect(url_for('receitas'))

@app.route('/receitas/editar-fixa/<int:id>', methods=['POST'])
@login_required
def editar_receita_fixa(id):
    receita = ReceitaPessoal.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    receita.descricao = request.form['descricao']
    receita.valor = float(request.form['valor'])
    db.session.commit()
    return redirect(url_for('receitas'))

@app.route('/receitas/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_receita(id):
    receita = ReceitaPessoal.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    db.session.delete(receita)
    db.session.commit()
    return redirect(url_for('receitas'))

# IMPORTAR
@app.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    if request.method == 'GET':
        return render_template('importar.html')
    arquivo = request.files['arquivo']
    if not arquivo:
        return redirect(url_for('importar'))
    wb = openpyxl.load_workbook(arquivo)
    resultado = {'clientes': 0, 'despesas': 0, 'erros': []}
    uid = current_user.id
    if 'RECEITAS' in wb.sheetnames:
        ws = wb['RECEITAS']
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not row[0]:
                continue
            nome = str(row[0]).strip()
            if not nome:
                continue
            try:
                dia = int(row[1]) if row[1] else 1
                valor = float(row[2]) if row[2] else 0
                existente = Cliente.query.filter_by(nome=nome, usuario_id=uid).first()
                if not existente:
                    cliente = Cliente(nome=nome, valor_mensalidade=valor, dia_vencimento=dia, status='ativo', usuario_id=uid)
                    db.session.add(cliente)
                    resultado['clientes'] += 1
            except Exception as e:
                resultado['erros'].append(f"Cliente {nome}: {str(e)}")
    if 'DESPESAS' in wb.sheetnames:
        ws = wb['DESPESAS']
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        meses_colunas = []
        for i, cell in enumerate(header):
            if isinstance(cell, datetime):
                meses_colunas.append((i, cell))
        for row in rows[1:]:
            if not row[0]:
                continue
            nome = str(row[0]).strip()
            if not nome:
                continue
            tipo = str(row[1]).strip().lower() if row[1] else 'v'
            try:
                if tipo == 'f':
                    valor_fixo = 0
                    for col_idx, mes_data in meses_colunas:
                        val = row[col_idx]
                        if val and isinstance(val, (int, float)) and val > 0:
                            valor_fixo = float(val)
                            break
                    existente = DespesaFixa.query.filter_by(descricao=nome, usuario_id=uid).first()
                    if not existente:
                        despesa = DespesaFixa(descricao=nome, valor=valor_fixo, ativa=True, usuario_id=uid)
                        db.session.add(despesa)
                    resultado['despesas'] += 1
                else:
                    existente = DespesaVariavel.query.filter_by(descricao=nome, usuario_id=uid).first()
                    if not existente:
                        despesa = DespesaVariavel(descricao=nome, cor='#f1efe8', usuario_id=uid)
                        db.session.add(despesa)
                        db.session.flush()
                    else:
                        despesa = existente
                    for col_idx, mes_data in meses_colunas:
                        valor = row[col_idx]
                        if valor and isinstance(valor, (int, float)) and valor > 0:
                            mes = mes_data.month
                            ano = mes_data.year
                            lanc_existente = LancamentoDespesaVariavel.query.filter_by(
                                despesa_id=despesa.id, mes=mes, ano=ano
                            ).first()
                            if not lanc_existente:
                                lanc = LancamentoDespesaVariavel(
                                    despesa_id=despesa.id, mes=mes, ano=ano, valor=float(valor), pago=False
                                )
                                db.session.add(lanc)
                    resultado['despesas'] += 1
            except Exception as e:
                resultado['erros'].append(f"Despesa {nome}: {str(e)}")
    db.session.commit()
    return render_template('importar.html', resultado=resultado)

@app.route('/migrate-xyz123')
def migrate():
    with db.engine.connect() as conn:
        try:
            conn.execute(db.text('ALTER TABLE despesa_fixa ADD COLUMN dia_vencimento INTEGER'))
            conn.commit()
            return 'Migração concluída!'
        except Exception as e:
            return f'Erro (pode já existir): {str(e)}'

#@app.route('/recreate-db-xyz123')
#def recreate_db():
#    db.drop_all()
#    db.create_all()
#    return 'Banco recriado!'

#@app.route('/admin-setup-xyz123')
#def admin_setup():
#    usuarios = Usuario.query.all()
#    resultado = '<h2>Usuários:</h2>'
#    for u in usuarios:
#        resultado += f'<p>ID: {u.id} | Email: {u.email} | Admin: {u.admin}</p>'
#    return resultado

#@app.route('/admin-set-xyz123/<email>')
#def admin_set(email):
 #   u = Usuario.query.filter_by(email=email).first()
  #  if u:
   #     u.admin = True
    #    db.session.commit()
     #   return f'Admin definido para {u.email}!'
    #return 'Usuário não encontrado.'

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)